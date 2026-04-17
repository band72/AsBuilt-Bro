import csv
import os
from PIL import Image, ImageDraw
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment, PatternFill

def get_color(code, desc):
    sys = code.upper()
    d = desc.upper()
    if 'CHIL' in d or 'CH' in sys: return '#87CEFA' # lightskyblue
    if 'WASTE' in d or 'SEW' in d or 'WW' in sys: return '#008000' # green
    if 'WATER' in d or sys == 'W' or 'WAT' in sys: return '#0000FF' # blue
    if 'STORM' in d or 'DRAIN' in d or 'ST' in sys or sys == 'D': return '#00FFFF' # cyan
    if 'RECLAIM' in d or 'REC' in sys or 'R' in sys: return '#800080' # purple
    if 'GAS' in d or 'G' == sys: return '#FFA500' # orange
    if 'ELEC' in d or sys == 'E' or 'EL' in sys: return '#FF0000' # red
    return '#808080' # gray

def get_symbol_type(desc):
    d = desc.upper()
    if 'MANHOLE' in d or 'VAULT' in d or 'CB' in d or 'INLET' in d or 'CATCH BASIN' in d: return 'manhole'
    if 'VALVE' in d: return 'valve'
    if 'FITTING' in d: return 'fitting'
    if 'HYDRANT' in d: return 'hydrant'
    if 'METER' in d: return 'meter'
    if 'POINT' in d: return 'point'
    if 'RUN' in d or 'PIPE' in d: return 'line'
    return 'default'

def draw_symbol(sym_type, color, filename):
    img = Image.new('RGBA', (60, 60), (0, 0, 0, 0)) # transparent background
    d = ImageDraw.Draw(img)
    
    if sym_type == 'manhole':
        d.ellipse([10, 10, 50, 50], fill=color, outline='white', width=2)
        d.line([20, 20, 40, 40], fill='white', width=2)
        d.line([20, 40, 40, 20], fill='white', width=2)
    elif sym_type == 'valve':
        d.polygon([(10, 20), (30, 30), (10, 40)], fill=color, outline='white', width=2)
        d.polygon([(50, 20), (30, 30), (50, 40)], fill=color, outline='white', width=2)
    elif sym_type == 'fitting':
        d.ellipse([20, 20, 40, 40], fill=color, outline='white', width=2)
    elif sym_type == 'hydrant':
        d.ellipse([20, 20, 40, 40], fill=color, outline='white', width=2)
        d.line([30, 10, 30, 50], fill='white', width=2)
        d.line([10, 30, 50, 30], fill='white', width=2)
    elif sym_type == 'meter':
        d.rectangle([15, 20, 45, 40], fill=color, outline='white', width=2)
    elif sym_type == 'line':
        d.line([5, 30, 55, 30], fill=color, width=4)
    else:
        # Default Point
        d.polygon([(30, 10), (50, 40), (10, 40)], fill=color, outline='white', width=2)
        
    img.save(filename)

def main():
    os.makedirs('symbols_library', exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Utility Symbols"
    
    # Fill background to dark so white outlines show up
    # Wait, in Excel it's usually white, let's just make the cells dark
    for row in range(1, 100):
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
            ws.cell(row=row, column=col).font = Font(color="FFFFFF")
            
    headers = ["Local Code", "System Code", "Description", "Symbol Preview"]
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    
    row_num = 2
    with open('MasterUtilityCodes.csv', 'r') as f:
        reader = csv.reader(f)
        try: next(reader)
        except: pass
        
        for row in reader:
            if not row or len(row) < 3: continue
            local, sys, desc = row[0], row[1], row[2]
            
            c_local = ws.cell(row=row_num, column=1, value=local)
            c_sys = ws.cell(row=row_num, column=2, value=sys)
            c_desc = ws.cell(row=row_num, column=3, value=desc)
            
            for c in [c_local, c_sys, c_desc]:
                c.alignment = Alignment(vertical='center')
            
            color = get_color(local, desc)
            sym_type = get_symbol_type(desc)
            
            filename = os.path.join('symbols_library', f"{local}_{sys}.png").replace("/", "_").replace("\\", "_")
            if not os.path.exists(filename):
                draw_symbol(sym_type, color, filename)
            
            img = OpenpyxlImage(filename)
            img.width, img.height = 24, 24
            
            # center image slightly by placing it in cell
            ws.row_dimensions[row_num].height = 30
            ws.add_image(img, f"D{row_num}")
            
            row_num += 1

    wb.save('UtilitySymbols_Mapping.xlsx')
    print("Excel mapping created at UtilitySymbols_Mapping.xlsx")

if __name__ == '__main__':
    main()
