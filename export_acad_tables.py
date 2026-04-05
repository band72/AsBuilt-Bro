import pandas as pd
import openpyxl
import os

BASE_DIR  = r"C:\Users\Daryl Banks\Downloads\70498-S1A_TEST FILE - Standard (1)"
TABLE_DIR = os.path.join(BASE_DIR, "table-ssmc")
TEMPLATE  = os.path.join(BASE_DIR, "Segment1A_All_Tables_Database_Ready.xlsx")

# ────────────────────────────────────────────────────────────────────
# Source column names (from our extracted CSVs)
# ────────────────────────────────────────────────────────────────────
FITTINGS_SRC_COLS = [
    "Fitting Number:",          # C1  -> PartKey
    "Subtype",                  # C2  -> Subtype
    "Facility Owner",           # C3  -> FacilityOwner
    "Primary Fitting Size (inches) ",  # C4 -> Size
    "Secondary Fitting Size (inches)", # C5 -> SizeSecondary
    "Material",                 # C6  -> Material (skip Manufacturer col)
    "Final Grade Elevation (feet)",    # C7  -> GradeElevation
    "Y Coordinates",            # C8  -> Northing
    "X Coordinates",            # C9  -> Easting
    "Latitude",                 # C10 -> Latitude
    "Longitude",                # C11 -> Longitude
]

VALVE_SRC_COLS = [
    "Valve Number",             # C1  -> PartKey
    "Subtype",                  # C2  -> Subtype
    "Valve Type",               # C3  -> ValveType
    "Facility Owner",           # C4  -> FacilityOwner
    "Valve Size (inches)",      # C5  -> Size
    "Valve Orientation",        # C6  -> Orientation
    "Open Direction",           # C7  -> OpenDirection
    "Turns To Open",            # C8  -> TurnsToOpen
    "Nut Elevation (feet)",     # C9  -> NutElevation
    "Finshed Grade Elevation (feet)",  # C10 -> GradeElevation
    "Depth To Nut (feet)",      # C11 -> DepthToNut
    "Manufacturer",             # C12 -> Manufacturer
    "X Coordinates",            # C13 -> Easting
    "Y Coordinates",            # C14 -> Northing
    "Latitude",                 # C15 -> Latitude
    "Longitude",                # C16 -> Longitude
]

WIREBOX_SRC_COLS = [
    "Fitting Number:",          # C1  -> PartKey
    "Subtype",                  # C2  -> Subtype
    "X Coordinates",            # C3  -> Easting
    "Y Coordinates",            # C4  -> Northing
    "Latitude",                 # C5  -> Latitude
    "Longitude",                # C6  -> Longitude
]

TOPOFPIPE_SRC_COLS = [
    "Point Number:",            # C1  -> PartKey
    "Pipe Location",            # C2  -> PipeRole
    "Subtype",                  # C3  -> Subtype
    "Facility Owner",           # C4  -> FacilityOwner
    "Primary Fitting  Size (inches) ", # C5 -> Size
    "Pipe Orientation",         # C6  -> Orientation
    "Pipe Class",               # C7  -> PipeClass
    "Pipe  Manufacturer",       # C8  -> Manufacturer
    "Pipe Material",            # C9  -> Material
    "Pipe Lining  Material",    # C10 -> LiningMaterial
    "Pipe Lining  Manufacturer",# C11 -> LiningManufacturer
    "Final Grade  Elevation (feet)", # C12 -> GradeElevation
    "Elevation  (feet)",        # C13 -> TopElevation
    "Cover (feet)",             # C14 -> Cover
    "X Coordinates",            # C15 -> Easting
    "Y Coordinates",            # C16 -> Northing
    "Latitude",                 # C17 -> Latitude
    "Longitude",                # C18 -> Longitude
]

IMPORTS = [
    {"src": "Water_Main_Fittings.xlsx",  "sheet": "Water Fitting",          "cols": FITTINGS_SRC_COLS},
    {"src": "Water_Main_Valves.xlsx",    "sheet": "Water Valve",            "cols": VALVE_SRC_COLS},
    {"src": "Water_Wire_Box.xlsx",       "sheet": "Water Locate Box",       "cols": WIREBOX_SRC_COLS},
    {"src": "Water_Main_TopOfPipe.xlsx", "sheet": "Water_Main_Top_Of_Pipe", "cols": TOPOFPIPE_SRC_COLS},
]

def remap_df(df, src_cols):
    """Pick and reorder columns by fuzzy match on stripped names."""
    # Build mapping of stripped col name -> actual df col name
    df_col_map = {c.strip(): c for c in df.columns}
    
    result_cols = []
    for wanted in src_cols:
        w = wanted.strip()
        if w in df_col_map:
            result_cols.append(df_col_map[w])
        else:
            # Try partial match
            match = next((c for c in df.columns if wanted.strip().lower() in c.strip().lower() or c.strip().lower() in wanted.strip().lower()), None)
            if match:
                result_cols.append(match)
                print(f"  Fuzzy matched '{wanted}' -> '{match}'")
            else:
                print(f"  WARNING: Column not found: '{wanted}' — inserting blank column")
                df[f"__BLANK_{w}"] = None
                result_cols.append(f"__BLANK_{w}")
    
    return df[result_cols]

print(f"Loading template: {TEMPLATE}")
wb = openpyxl.load_workbook(TEMPLATE)
print(f"Sheets: {wb.sheetnames}\n")

for imp in IMPORTS:
    src_path     = os.path.join(TABLE_DIR, imp["src"])
    target_sheet = imp["sheet"]

    if not os.path.exists(src_path):
        print(f"WARNING: Not found: {src_path}")
        continue

    df_raw = pd.read_excel(src_path)
    print(f"'{imp['src']}' raw cols: {list(df_raw.columns)}")
    
    df = remap_df(df_raw, imp["cols"])
    print(f"  Remapped to {df.shape[1]} cols for sheet '{target_sheet}'")

    if target_sheet in wb.sheetnames:
        ws = wb[target_sheet]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    else:
        ws = wb.create_sheet(title=target_sheet)
        # Write header row using positional numbers (service reads by col index, not name)
        for ci in range(1, df.shape[1] + 1):
            ws.cell(row=1, column=ci, value=f"Col{ci}")

    # Write data rows starting at row 2
    for ri, row in enumerate(df.itertuples(index=False), start=2):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=val)

    print(f"  Written {df.shape[0]} rows.\n")

wb.save(TEMPLATE)
print(f"SUCCESS! Saved: {TEMPLATE}")
